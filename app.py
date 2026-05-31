import os
from dotenv import load_dotenv
load_dotenv()
import midtransclient
from flask import Flask, render_template, request, jsonify, send_file, redirect, url_for
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from flask_bcrypt import Bcrypt
from datetime import datetime
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
import json, io, requests as req
import jwt as pyjwt
from functools import wraps
from routes.sync import sync_bp
from extensions import db

app = Flask(__name__)
app.config['SQLALCHEMY_DATABASE_URI'] = os.environ.get('DATABASE_URL', 'postgresql://martin:umkm1234@localhost/umkmpro')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SECRET_KEY'] = 'umkmpro-secret-2024'
db.init_app(app)
app.register_blueprint(sync_bp)
bcrypt = Bcrypt(app)
login_manager = LoginManager(app)
login_manager.login_view = 'login'

MIDTRANS_SERVER_KEY = os.environ.get('MIDTRANS_SERVER_KEY', '')
MIDTRANS_CLIENT_KEY = os.environ.get('MIDTRANS_CLIENT_KEY', '')
JWT_SECRET = os.environ.get('JWT_SECRET', 'umkmpro-secret-key')

def token_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        token = request.headers.get('Authorization', '').replace('Bearer ', '')
        if not token:
            # fallback ke session cookie (untuk web browser biasa)
            if current_user.is_authenticated:
                return f(*args, **kwargs)
            return jsonify({'error': 'Token required'}), 401
        try:
            data = pyjwt.decode(token, JWT_SECRET, algorithms=['HS256'])
            user = User.query.get(data['user_id'])
            if not user:
                return jsonify({'error': 'User not found'}), 401
        except pyjwt.ExpiredSignatureError:
            return jsonify({'error': 'Token expired'}), 401
        except Exception:
            return jsonify({'error': 'Token invalid'}), 401
        return f(*args, **kwargs)
    return decorated

print(f"Server Key loaded: {MIDTRANS_SERVER_KEY[:10]}...")

snap = midtransclient.Snap(
    is_production=False,
    server_key=MIDTRANS_SERVER_KEY
)

@app.after_request
def add_security_headers(response):
    response.headers['Content-Security-Policy'] = (
        "default-src *; "
        "script-src * 'unsafe-inline' 'unsafe-eval'; "
        "style-src * 'unsafe-inline'; "
        "frame-src *; "
        "connect-src *; "
        "img-src * data:;"
    )
    return response

class Toko(db.Model):
    __tablename__ = 'toko'
    id = db.Column(db.Integer, primary_key=True)
    nama = db.Column(db.String(200), default='Toko Saya')
    alamat = db.Column(db.String(500), default='')
    telepon = db.Column(db.String(50), default='')
    email = db.Column(db.String(100), default='')
    tagline = db.Column(db.String(200), default='')
    ppn_persen = db.Column(db.Integer, default=10)
    dibuat = db.Column(db.DateTime, default=datetime.now)
    fonnte_token = db.Column(db.String(200), default='')
    

class User(UserMixin, db.Model):
    __tablename__ = 'users'
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password = db.Column(db.String(200), nullable=False)
    role = db.Column(db.String(20), default='kasir')
    nama = db.Column(db.String(100))
    dibuat = db.Column(db.DateTime, default=datetime.now)

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

class Produk(db.Model):
    __tablename__ = 'produk'
    id = db.Column(db.String, primary_key=True)
    nama = db.Column(db.String(200), nullable=False)
    harga = db.Column(db.Integer, nullable=False)
    modal = db.Column(db.Integer, default=0)
    stok = db.Column(db.Integer, default=0)
    kategori = db.Column(db.String(100), default='Umum')
    satuan = db.Column(db.String(50), default='pcs')
    dibuat = db.Column(db.DateTime, default=datetime.now)
    updated_at = db.Column(db.DateTime, default=datetime.now)
    deleted_at = db.Column(db.DateTime, nullable=True)
    device_id  = db.Column(db.String, nullable=True)

class Transaksi(db.Model):
    __tablename__ = 'transaksi'
    id = db.Column(db.String, primary_key=True)
    pelanggan = db.Column(db.String(200), default='Umum')
    kasir = db.Column(db.String(100), default='Kasir')
    subtotal = db.Column(db.Integer, default=0)
    ppn = db.Column(db.Integer, default=0)
    total = db.Column(db.Integer, nullable=False)
    items = db.Column(db.Text)
    tanggal = db.Column(db.DateTime, default=datetime.now)
    updated_at = db.Column(db.DateTime, default=datetime.now)
    deleted_at = db.Column(db.DateTime, nullable=True)
    device_id  = db.Column(db.String, nullable=True)

class Jurnal(db.Model):
    __tablename__ = "jurnal"
    id = db.Column(db.Integer, primary_key=True)
    tanggal = db.Column(db.DateTime, default=datetime.now)
    keterangan = db.Column(db.String(500), nullable=False)
    kategori = db.Column(db.String(100), default="Umum")
    jenis = db.Column(db.String(20), default="pemasukan")
    debit = db.Column(db.Integer, default=0)
    kredit = db.Column(db.Integer, default=0)
    sumber = db.Column(db.String(50), default="manual")
    referensi_id = db.Column(db.Integer, default=0)
    dibuat_oleh = db.Column(db.String(100), default="")

class Invoice(db.Model):
    __tablename__ = 'invoice'
    id = db.Column(db.Integer, primary_key=True)
    nomor = db.Column(db.String(50))
    pelanggan = db.Column(db.String(200), nullable=False)
    produk = db.Column(db.String(200))
    jumlah = db.Column(db.Integer, default=1)
    harga = db.Column(db.Integer, default=0)
    total = db.Column(db.Integer, default=0)
    jatuh_tempo = db.Column(db.String(50))
    status = db.Column(db.String(50), default='Belum Bayar')
    tanggal = db.Column(db.DateTime, default=datetime.now)
    items = db.Column(db.Text, default='[]')

def kirim_whatsapp(nomor, pesan):
    # Cek token dari database toko dulu, kalau tidak ada pakai env variable
    toko = Toko.query.first()
    token = (toko.fonnte_token if toko and toko.fonnte_token else '') or os.environ.get('FONNTE_TOKEN', '')
    if not token:
        print("FONNTE_TOKEN tidak ditemukan")
        return False
    try:
        response = req.post('https://api.fonnte.com/send',
            headers={'Authorization': token},
            data={'target': nomor, 'message': pesan, 'countryCode': '62'}
        )
        print(f"WhatsApp sent: {response.json()}")
        return True
    except Exception as e:
        print(f"WhatsApp error: {str(e)}")
        return False

@app.route('/privacy')
def privacy():
    return render_template('privacy.html')

@app.route('/setup', methods=['GET', 'POST'])
def setup():
    if User.query.filter_by(role='owner').count() > 0:
        return redirect(url_for('login'))
    if request.method == 'POST':
        data = request.json
        toko = Toko.query.first()
        if not toko:
            toko = Toko()
            db.session.add(toko)
        toko.nama = data.get('nama_toko', 'Toko Saya')
        toko.alamat = data.get('alamat', '')
        toko.telepon = data.get('telepon', '')
        toko.email = data.get('email', '')
        owner = User(
            username=data.get('username', 'owner'),
            password=bcrypt.generate_password_hash(data.get('password', 'owner123')).decode('utf-8'),
            role='owner',
            nama=data.get('nama_owner', 'Owner')
        )
        db.session.add(owner)
        db.session.commit()
        return jsonify({'ok': True})
    return render_template('setup.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if User.query.filter_by(role='owner').count() == 0:
        return redirect(url_for('setup'))
    if request.method == 'POST':
        data = request.json
        user = User.query.filter_by(username=data['username']).first()
        if user and bcrypt.check_password_hash(user.password, data['password']):
            login_user(user)
            import datetime as dt
            token = pyjwt.encode({
                'user_id': user.id,
                'exp': dt.datetime.utcnow() + dt.timedelta(days=30)
            }, JWT_SECRET, algorithm='HS256')
            return jsonify({'ok': True, 'role': user.role, 'nama': user.nama, 'token': token})
        return jsonify({'ok': False, 'pesan': 'Username atau password salah'}), 401
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

@app.route('/api/me')
@login_required
def me():
    return jsonify({'id': current_user.id, 'username': current_user.username, 'role': current_user.role, 'nama': current_user.nama})

@app.route('/')
@login_required
def index():
    return render_template('index.html')

@app.route('/api/toko', methods=['GET'])
@login_required
def get_toko():
    toko = Toko.query.first()
    if not toko:
        toko = Toko()
        db.session.add(toko)
        db.session.commit()
    return jsonify({'nama': toko.nama, 'alamat': toko.alamat, 'telepon': toko.telepon, 'email': toko.email, 'tagline': toko.tagline, 'ppn_persen': toko.ppn_persen or 10, 'fonnte_token': toko.fonnte_token or ''})

@app.route('/api/toko', methods=['POST'])
@login_required
def update_toko():
    if current_user.role != 'owner':
        return jsonify({'error': 'Akses ditolak'}), 403
    toko = Toko.query.first()
    if not toko:
        toko = Toko()
        db.session.add(toko)
    b = request.json
    toko.nama = b.get('nama', toko.nama)
    toko.alamat = b.get('alamat', toko.alamat)
    toko.telepon = b.get('telepon', toko.telepon)
    toko.email = b.get('email', toko.email)
    toko.tagline = b.get('tagline', toko.tagline)
    toko.ppn_persen = int(b.get('ppn_persen', toko.ppn_persen or 10))
    toko.fonnte_token = b.get('fonnte_token', toko.fonnte_token or '')
    db.session.commit()
    return jsonify({'ok': True})

@app.route('/api/produk', methods=['GET'])
@token_required
def get_produk():
    produk = Produk.query.all()
    return jsonify([{'id': p.id, 'nama': p.nama, 'harga': p.harga, 'modal': p.modal, 'stok': p.stok, 'kategori': p.kategori, 'satuan': p.satuan or 'pcs'} for p in produk])

@app.route('/api/produk', methods=['POST'])
@token_required
def add_produk():
    b = request.json
    p = Produk(nama=b['nama'], harga=int(b['harga']), modal=int(b.get('modal', 0)), stok=int(b.get('stok', 99)), kategori=b.get('kategori', 'Umum'), satuan=b.get('satuan', 'pcs'))
    db.session.add(p)
    db.session.commit()
    return jsonify({'id': p.id, 'nama': p.nama, 'harga': p.harga, 'modal': p.modal, 'stok': p.stok, 'satuan': p.satuan})

@app.route('/api/produk/<int:pid>/stok', methods=['POST'])
@token_required
def update_stok(pid):
    p = Produk.query.get_or_404(pid)
    b = request.json
    p.stok = int(b.get('stok', p.stok))
    db.session.commit()
    return jsonify({'ok': True, 'stok': p.stok})
    
@app.route('/api/produk/<int:pid>', methods=['DELETE'])
@token_required
def del_produk(pid):
    p = Produk.query.get_or_404(pid)
    db.session.delete(p)
    db.session.commit()
    return jsonify({'ok': True})

@app.route('/api/transaksi', methods=['GET'])
@token_required
def get_transaksi():
    transaksi = Transaksi.query.order_by(Transaksi.tanggal.desc()).all()
    total_subtotal = sum(t.subtotal if t.subtotal is not None else t.total for t in transaksi)
    total_ppn = sum(t.ppn or 0 for t in transaksi)
    total_semua = sum(t.total for t in transaksi)
    return jsonify({
        'transaksi': [{'id': t.id, 'pelanggan': t.pelanggan, 'kasir': t.kasir, 'subtotal': t.subtotal if t.subtotal is not None else t.total, 'ppn': t.ppn or 0, 'total': t.total, 'items': json.loads(t.items or '[]'), 'tanggal': t.tanggal.strftime('%d/%m/%Y %H:%M')} for t in transaksi],
        'total_subtotal': total_subtotal,
        'total_ppn': total_ppn,
        'total_semua': total_semua
    })

@app.route('/api/transaksi', methods=['POST'])
@token_required
def add_transaksi():
    b = request.json
    items = b['items']
    total = 0
    for item in items:
        p = Produk.query.get(item['id'])
        if p:
            p.stok = max(0, p.stok - item['qty'])
            total += p.harga * item['qty']
    toko = Toko.query.first()
    ppn_persen = toko.ppn_persen if toko and toko.ppn_persen is not None else 10
    ppn = round(total * ppn_persen / 100)
    total_ppn = total + ppn
    t = Transaksi(
        pelanggan=b.get('pelanggan', 'Umum'),
        kasir=current_user.nama or current_user.username,
        subtotal=total,
        ppn=ppn,
        total=total_ppn,
        items=json.dumps(items)
    )
    db.session.add(t)
    db.session.commit()
    nomor = b.get('nomor_wa', '')
    if nomor:
        nama_toko = toko.nama if toko else 'UMKM Pro'
        ppn_persen_wa = toko.ppn_persen if toko and toko.ppn_persen is not None else 10
        ppn_wa = round(total * ppn_persen_wa / 100)
        total_ppn_wa = total + ppn_wa
        pesan = f"✅ *Struk Pembelian {nama_toko}*\n\n"
        pesan += f"No: TRX-{t.id:04d}\n"
        pesan += f"Pelanggan: {t.pelanggan}\n"
        pesan += f"Kasir: {t.kasir}\n"
        pesan += f"Tanggal: {t.tanggal.strftime('%d/%m/%Y %H:%M')}\n\n"
        for item in items:
            p = Produk.query.get(item['id'])
            if p:
                pesan += f"- {p.nama} x{item['qty']} = Rp {p.harga * item['qty']:,}\n"
        pesan += f"\nSubtotal: Rp {total:,}\n"
        pesan += f"PPN {ppn_persen_wa}%: Rp {ppn_wa:,}\n"
        pesan += f"*Total: Rp {total_ppn_wa:,}*\n\n"
        pesan += "Terima kasih telah berbelanja! 🙏"
        kirim_whatsapp(nomor, pesan)
    try:
        j = Jurnal(
            keterangan='Penjualan - ' + str(t.pelanggan),
            kategori='Penjualan',
            jenis='pemasukan',
            debit=total,
            kredit=0,
            sumber='kasir',
            referensi_id=t.id,
            dibuat_oleh=current_user.nama or current_user.username
        )
        db.session.add(j)
        db.session.commit()
    except Exception as e:
        print('Jurnal error: ' + str(e))
    return jsonify({'id': t.id, 'total': total})

@app.route('/api/invoice', methods=['GET'])
@token_required
def get_invoice():
    invoices = Invoice.query.order_by(Invoice.tanggal.desc()).all()
    return jsonify([{'id': i.id, 'nomor': i.nomor, 'pelanggan': i.pelanggan, 'produk': i.produk, 'jumlah': i.jumlah, 'harga': i.harga, 'total': i.total, 'jatuh_tempo': i.jatuh_tempo, 'status': i.status, 'tanggal': i.tanggal.strftime('%d/%m/%Y'), 'items': json.loads(i.items or '[]')} for i in invoices])

@app.route('/api/invoice', methods=['POST'])
@token_required
def add_invoice():
    b = request.json
    nomor = f"INV-{datetime.now().strftime('%Y%m%d')}-{Invoice.query.count()+1:03d}"
    items = b.get('items', [])
    total = sum(item['jumlah'] * item['harga'] for item in items)
    # Ambil nama produk pertama untuk kolom produk (backward compat)
    produk_nama = items[0]['nama'] if items else b.get('produk', '')
    i = Invoice(
        nomor=nomor,
        pelanggan=b['pelanggan'],
        produk=produk_nama,
        jumlah=len(items),
        harga=total,
        total=total,
        items=json.dumps(items),
        jatuh_tempo=b.get('jatuh_tempo', '-')
    )
    db.session.add(i)
    db.session.commit()
    return jsonify({'id': i.id, 'nomor': i.nomor})

@app.route('/api/invoice/<int:iid>/lunas', methods=['POST'])
@token_required
def lunas_invoice(iid):
    i = Invoice.query.get_or_404(iid)
    i.status = 'Lunas'
    db.session.commit()
    return jsonify({'ok': True})

@app.route('/api/invoice/<int:iid>/pdf')
@token_required
def cetak_invoice(iid):
    i = Invoice.query.get_or_404(iid)
    toko = Toko.query.first()
    nama_toko = toko.nama if toko else 'UMKM Pro'
    alamat_toko = toko.alamat if toko else ''
    telepon_toko = toko.telepon if toko else ''
    email_toko = toko.email if toko else ''
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
        rightMargin=2*cm, leftMargin=2*cm,
        topMargin=2*cm, bottomMargin=2*cm)
    from reportlab.platypus import HRFlowable
    elemen = []

    # Styles
    s_judul = ParagraphStyle('judul', fontSize=16, fontName='Helvetica-Bold', alignment=1, spaceAfter=16)
    s_bold = ParagraphStyle('bold', fontSize=10, fontName='Helvetica-Bold', spaceAfter=2)
    s_normal = ParagraphStyle('normal', fontSize=10, fontName='Helvetica', spaceAfter=2)
    s_small = ParagraphStyle('small', fontSize=9, fontName='Helvetica', textColor=colors.HexColor('#444444'), spaceAfter=2)
    s_kanan = ParagraphStyle('kanan', fontSize=10, fontName='Helvetica', alignment=2, spaceAfter=2)
    s_kanan_bold = ParagraphStyle('kanan_bold', fontSize=12, fontName='Helvetica-Bold', alignment=2, spaceAfter=4)

    # JUDUL
    elemen.append(Paragraph('INVOICE', s_judul))
    elemen.append(HRFlowable(width='100%', thickness=1, color=colors.black, spaceAfter=12))

    # Header: Info toko kiri, Info invoice kanan
    header_data = [[
        Paragraph(f'<b>{nama_toko}</b>', s_bold),
        Paragraph(f'No. {i.nomor}', s_normal)
    ],[
        Paragraph(alamat_toko, s_small),
        Paragraph(f'Tanggal: {i.tanggal.strftime("%d/%m/%Y")}', s_normal)
    ],[
        Paragraph(f'Telp: {telepon_toko}' if telepon_toko else '', s_small),
        Paragraph(f'Jatuh Tempo: {i.jatuh_tempo}', s_normal)
    ],[
        Paragraph(email_toko, s_small),
        Paragraph('', s_normal)
    ]]
    header_tabel = Table(header_data, colWidths=[10*cm, 7*cm])
    header_tabel.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('ALIGN', (1,0), (1,-1), 'RIGHT'),
    ]))
    elemen.append(header_tabel)
    elemen.append(Spacer(1, 0.5*cm))
    elemen.append(HRFlowable(width='100%', thickness=0.5, color=colors.HexColor('#cccccc'), spaceAfter=10))

    # Ditujukan kepada
    elemen.append(Paragraph('<b>Ditujukan Kepada:</b>', s_bold))
    elemen.append(Paragraph(i.pelanggan, s_normal))
    elemen.append(Spacer(1, 0.6*cm))

    # Tabel produk
    data = [['No.', 'Deskripsi', 'Jumlah', 'Satuan', 'Harga Satuan', 'Total']]
    items_list = json.loads(i.items or '[]')
    if items_list:
        for idx, item in enumerate(items_list):
            subtotal = item['jumlah'] * item['harga']
            p_item = Produk.query.get(item.get('id', 0))
            satuan_item = (p_item.satuan or 'pcs') if p_item else item.get('satuan', 'pcs')
            data.append([str(idx+1), item['nama'], str(item['jumlah']), satuan_item, f"Rp {item['harga']:,}", f"Rp {subtotal:,}"])
    else:
        data.append(['1', i.produk, str(i.jumlah), 'pcs', f"Rp {i.harga:,}", f"Rp {i.total:,}"])
    tabel = Table(data, colWidths=[1*cm, 6*cm, 2*cm, 2*cm, 3.5*cm, 3.5*cm])
    tabel.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#4472C4')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 10),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('ALIGN', (1,0), (1,-1), 'LEFT'),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#cccccc')),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#f5f5f5')]),
        ('TOPPADDING', (0,0), (-1,-1), 8),
        ('BOTTOMPADDING', (0,0), (-1,-1), 8),
        ('LEFTPADDING', (0,0), (-1,-1), 8),
    ]))
    elemen.append(tabel)
    elemen.append(Spacer(1, 0.3*cm))

    # Total (rata kanan)
    subtotal_inv = sum(item["jumlah"] * item["harga"] for item in items_list) if items_list else i.harga
    toko_ppn = Toko.query.first()
    ppn_persen_inv = toko_ppn.ppn_persen if toko_ppn and toko_ppn.ppn_persen is not None else 10
    ppn_inv = round(subtotal_inv * ppn_persen_inv / 100)
    total_inv = subtotal_inv + ppn_inv
    total_data = [
        ['', 'Subtotal:', f"Rp {subtotal_inv:,}"],
        ['', f'PPN {ppn_persen_inv}%:', f"Rp {ppn_inv:,}"],
        ['', 'Total:', f"Rp {total_inv:,}"],
    ]
    total_tabel = Table(total_data, colWidths=[11*cm, 3*cm, 4*cm])
    total_tabel.setStyle(TableStyle([
        ('ALIGN', (1,0), (-1,-1), 'RIGHT'),
        ('FONTNAME', (0,0), (-1,-1), 'Helvetica'),
        ('FONTSIZE', (0,0), (-1,-1), 10),
        ('TOPPADDING', (0,0), (-1,-1), 6),
    ]))
    elemen.append(total_tabel)
    elemen.append(Spacer(1, 0.5*cm))
    elemen.append(HRFlowable(width='100%', thickness=0.5, color=colors.HexColor('#cccccc'), spaceAfter=10))

    # Keterangan
    elemen.append(Paragraph('<b>Keterangan:</b>', s_bold))
    status_text = 'Lunas' if i.status == 'Lunas' else 'Belum Bayar'
    elemen.append(Paragraph(f'Status pembayaran: {status_text}', s_normal))
    elemen.append(Spacer(1, 1.5*cm))

    # Tanda tangan
    ttd_data = [[
        Paragraph('', s_normal),
        Paragraph(f'( {current_user.nama} )', s_kanan_bold)
    ],[
        Paragraph('', s_normal),
        Paragraph('Dibuat oleh', s_kanan)
    ]]
    ttd_tabel = Table(ttd_data, colWidths=[10*cm, 7*cm])
    ttd_tabel.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'BOTTOM'),
    ]))
    elemen.append(ttd_tabel)

    doc.build(elemen)
    buffer.seek(0)
    return send_file(buffer, mimetype='application/pdf', download_name=f'invoice-{i.nomor}.pdf')

@app.route('/api/struk/<int:tid>')
@login_required
def cetak_struk(tid):
    t = Transaksi.query.get_or_404(tid)
    items = json.loads(t.items or '[]')
    toko = Toko.query.first()
    nama_toko = toko.nama if toko else 'UMKM Pro'
    alamat_toko = toko.alamat if toko else ''
    telepon_toko = toko.telepon if toko else ''
    email_toko = toko.email if toko else ''
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
        rightMargin=2*cm, leftMargin=2*cm,
        topMargin=2*cm, bottomMargin=2*cm)
    from reportlab.platypus import HRFlowable
    elemen = []

    # Styles
    s_judul = ParagraphStyle('judul', fontSize=16, fontName='Helvetica-Bold', alignment=1, spaceAfter=16)
    s_bold = ParagraphStyle('bold', fontSize=10, fontName='Helvetica-Bold', spaceAfter=2)
    s_normal = ParagraphStyle('normal', fontSize=10, fontName='Helvetica', spaceAfter=2)
    s_small = ParagraphStyle('small', fontSize=9, fontName='Helvetica', textColor=colors.HexColor('#666666'), spaceAfter=1)
    s_kanan = ParagraphStyle('kanan', fontSize=10, fontName='Helvetica', alignment=2, spaceAfter=2)
    s_kanan_bold = ParagraphStyle('kanan_bold', fontSize=12, fontName='Helvetica-Bold', alignment=2, spaceAfter=4)

    # JUDUL
    elemen.append(Paragraph('STRUK PEMBELIAN', s_judul))
    elemen.append(HRFlowable(width='100%', thickness=1, color=colors.HexColor('#1e3a8a'), spaceAfter=12))

    # Header: Info toko kiri, Info transaksi kanan
    header_data = [[
        Paragraph(f'<b>{nama_toko}</b>', s_bold),
        Paragraph(f'No. TRX-{t.id:04d}', s_normal)
    ],[
        Paragraph(alamat_toko, s_small) if alamat_toko else Paragraph('', s_small),
        Paragraph(f'Tanggal: {t.tanggal.strftime("%d/%m/%Y %H:%M")}', s_normal)
    ],[
        Paragraph(f'Telp: {telepon_toko}', s_small) if telepon_toko else Paragraph('', s_small),
        Paragraph(f'Kasir: {t.kasir}', s_normal)
    ],[
        Paragraph(email_toko, s_small) if email_toko else Paragraph('', s_small),
        Paragraph('', s_normal)
    ]]
    header_tabel = Table(header_data, colWidths=[10*cm, 7*cm])
    header_tabel.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('ALIGN', (1,0), (1,-1), 'RIGHT'),
    ]))
    elemen.append(header_tabel)
    elemen.append(Spacer(1, 0.5*cm))
    elemen.append(HRFlowable(width='100%', thickness=0.5, color=colors.HexColor('#cccccc'), spaceAfter=10))

    # Info pelanggan
    pelanggan_data = [[
        Paragraph('PELANGGAN:', s_small),
        Paragraph('', s_small)
    ],[
        Paragraph(t.pelanggan, s_bold),
        Paragraph('', s_normal)
    ]]
    pelanggan_tabel = Table(pelanggan_data, colWidths=[12*cm, 4*cm])
    pelanggan_tabel.setStyle(TableStyle([('VALIGN', (0,0), (-1,-1), 'TOP')]))
    elemen.append(pelanggan_tabel)
    elemen.append(Spacer(1, 0.5*cm))

    # Tabel produk
    data = [['No.', 'Produk', 'Qty', 'Satuan', 'Harga Satuan', 'Total']]
    for idx, item in enumerate(items):
        p = Produk.query.get(item['id'])
        if p:
            subtotal = p.harga * item['qty']
            data.append([str(idx+1), p.nama, str(item['qty']), p.satuan or 'pcs', f"Rp {p.harga:,}", f"Rp {subtotal:,}"])
    tabel = Table(data, colWidths=[1*cm, 5.5*cm, 1.5*cm, 2*cm, 3.5*cm, 3.5*cm])
    tabel.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1e3a8a')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 10),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('ALIGN', (1,0), (1,-1), 'LEFT'),
        ('ALIGN', (4,0), (-1,-1), 'RIGHT'),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#e2e8f0')),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#f8faff')]),
        ('TOPPADDING', (0,0), (-1,-1), 8),
        ('BOTTOMPADDING', (0,0), (-1,-1), 8),
        ('LEFTPADDING', (0,0), (-1,-1), 8),
    ]))
    elemen.append(tabel)
    elemen.append(Spacer(1, 0.3*cm))

    # Total
    subtotal = t.subtotal if t.subtotal else t.total
    toko_data = Toko.query.first()
    ppn_persen = toko_data.ppn_persen if toko_data and toko_data.ppn_persen is not None else 10
    ppn = t.ppn if t.ppn else 0
    total_data = [
        ['', 'Subtotal:', f"Rp {subtotal:,}"],
        ['', f'PPN {ppn_persen}%:', f"Rp {ppn:,}"],
        ['', 'TOTAL:', f"Rp {t.total:,}"]
    ]
    total_tabel = Table(total_data, colWidths=[10*cm, 3*cm, 4*cm])
    total_tabel.setStyle(TableStyle([
        ('ALIGN', (1,0), (-1,-1), 'RIGHT'),
        ('FONTNAME', (0,0), (-1,-1), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 11),
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('LINEABOVE', (0,0), (-1,0), 1.5, colors.HexColor('#1e3a8a')),
    ]))
    elemen.append(total_tabel)
    elemen.append(Spacer(1, 0.8*cm))
    elemen.append(HRFlowable(width='100%', thickness=0.5, color=colors.HexColor('#e2e8f0'), spaceAfter=8))

    # Footer
    footer_data = [[
        Paragraph('Terima kasih telah berbelanja!', s_small),
        Paragraph(f'{nama_toko} · {alamat_toko}', ParagraphStyle('footer', fontSize=8, fontName='Helvetica', textColor=colors.HexColor('#999999'), alignment=2))
    ]]
    footer_tabel = Table(footer_data, colWidths=[8*cm, 9*cm])
    footer_tabel.setStyle(TableStyle([('VALIGN', (0,0), (-1,-1), 'TOP')]))
    elemen.append(footer_tabel)

    doc.build(elemen)
    buffer.seek(0)
    return send_file(buffer, mimetype='application/pdf', download_name=f'struk-{t.id}.pdf')

@app.route('/api/bayar', methods=['POST'])
@login_required
def bayar():
    b = request.json
    items = b['items']
    total = 0
    item_details = []
    for item in items:
        p = Produk.query.get(item['id'])
        if p:
            subtotal = p.harga * item['qty']
            total += subtotal
            item_details.append({'id': str(p.id), 'price': p.harga, 'quantity': item['qty'], 'name': p.nama[:50]})
    order_id = f"UMKM-{datetime.now().strftime('%Y%m%d%H%M%S')}"
    param = {
        'transaction_details': {'order_id': order_id, 'gross_amount': total},
        'item_details': item_details,
        'customer_details': {'first_name': b.get('pelanggan', 'Pelanggan')}
    }
    try:
        transaction = snap.create_transaction(param)
        return jsonify({'ok': True, 'token': transaction['token'], 'redirect_url': transaction['redirect_url'], 'order_id': order_id, 'total': total})
    except Exception as e:
        print(f"Midtrans error: {str(e)}")
        return jsonify({'ok': False, 'error': str(e)}), 500

@app.route('/api/users', methods=['GET'])
@login_required
def get_users():
    if current_user.role != 'owner':
        return jsonify({'error': 'Akses ditolak'}), 403
    users = User.query.all()
    return jsonify([{'id': u.id, 'username': u.username, 'nama': u.nama, 'role': u.role} for u in users])

@app.route('/api/users', methods=['POST'])
@login_required
def add_user():
    if current_user.role != 'owner':
        return jsonify({'error': 'Akses ditolak'}), 403
    b = request.json
    if User.query.filter_by(username=b['username']).first():
        return jsonify({'error': 'Username sudah dipakai'}), 400
    u = User(username=b['username'], password=bcrypt.generate_password_hash(b['password']).decode('utf-8'), role=b.get('role', 'kasir'), nama=b['nama'])
    db.session.add(u)
    db.session.commit()
    return jsonify({'id': u.id, 'username': u.username, 'nama': u.nama, 'role': u.role})

@app.route('/api/users/<int:uid>', methods=['DELETE'])
@login_required
def del_user(uid):
    if current_user.role != 'owner':
        return jsonify({'error': 'Akses ditolak'}), 403
    if uid == current_user.id:
        return jsonify({'error': 'Tidak bisa hapus akun sendiri'}), 400
    u = User.query.get_or_404(uid)
    db.session.delete(u)
    db.session.commit()
    return jsonify({'ok': True})

@app.route('/api/produk/template')
@login_required
def download_template():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    import io
    
    wb = Workbook()
    ws = wb.active
    ws.title = 'Produk'
    
    header_fill = PatternFill(start_color='1e3a8a', end_color='1e3a8a', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    example_fill = PatternFill(start_color='dbeafe', end_color='dbeafe', fill_type='solid')
    
    headers = ['Nama Produk', 'Kategori', 'Harga Jual', 'Modal', 'Stok']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Contoh data
    contoh = [
        ['Beras Premium 5kg', 'Makanan', 85000, 65000, 50],
        ['Minyak Goreng 1L', 'Makanan', 22000, 18000, 30],
        ['Sabun Mandi', 'Kesehatan', 5000, 3500, 100],
    ]
    for row, data in enumerate(contoh, 2):
        for col, val in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill = example_fill
    
    # Kolom kategori dengan keterangan
    ws.cell(row=6, column=1, value='* Kategori yang tersedia:')
    ws.cell(row=7, column=1, value='Umum, Makanan, Minuman, Elektronik, Pakaian, Kesehatan, Lainnya')
    
    # Auto width
    for col in range(1, 6):
        ws.column_dimensions[get_column_letter(col)].width = 22
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return send_file(buffer, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', download_name='template-produk.xlsx')

@app.route('/api/produk/import', methods=['POST'])
@login_required
def import_produk():
    from openpyxl import load_workbook
    import io
    
    if 'file' not in request.files:
        return jsonify({'error': 'File tidak ditemukan'}), 400
    
    file = request.files['file']
    if not file.filename.endswith('.xlsx'):
        return jsonify({'error': 'File harus format .xlsx'}), 400
    
    try:
        wb = load_workbook(io.BytesIO(file.read()))
        ws = wb.active
        
        berhasil = 0
        gagal = 0
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            print(f"Row data: {row}")
            # Skip baris kosong dan baris keterangan
            if not row[0] or str(row[0]).startswith('*'):
                continue
            try:
                nama = str(row[0]).strip()
                kategori = str(row[1]).strip() if row[1] else 'Umum'
                harga = int(row[2]) if row[2] else 0
                modal = int(row[3]) if row[3] else 0
                stok = int(row[4]) if row[4] else 0
                
                if nama and harga > 0:
                    # Cek apakah produk sudah ada
                    existing = Produk.query.filter_by(nama=nama).first()
                    if existing:
                        existing.harga = harga
                        existing.modal = modal
                        existing.stok = stok
                        existing.kategori = kategori
                    else:
                        p = Produk(nama=nama, kategori=kategori, harga=harga, modal=modal, stok=stok)
                        db.session.add(p)
                    berhasil += 1
                else:
                    gagal += 1
            except:
                gagal += 1
        
        db.session.commit()
        return jsonify({'ok': True, 'berhasil': berhasil, 'gagal': gagal})
    except Exception as e:
        return jsonify({'error': str(e)}), 500
        
@app.route('/api/stok-alert')
@login_required
def stok_alert():
    batas = request.args.get('batas', 10, type=int)
    menipis = Produk.query.filter(Produk.stok <= batas).all()
    return jsonify({'jumlah': len(menipis), 'produk': [{'id': p.id, 'nama': p.nama, 'stok': p.stok, 'harga': p.harga} for p in menipis]})

@app.route('/api/laporan')
@token_required
def laporan():
    from collections import defaultdict
    transaksi = Transaksi.query.order_by(Transaksi.tanggal.asc()).all()
    per_hari = defaultdict(int)
    per_bulan = defaultdict(int)
    for t in transaksi:
        tgl = t.tanggal.strftime('%d/%m')
        per_hari[tgl] += t.total
        bln = t.tanggal.strftime('%b %Y')
        per_bulan[bln] += t.total
    produk_count = defaultdict(int)
    for t in transaksi:
        items = json.loads(t.items or '[]')
        for item in items:
            p = Produk.query.get(item['id'])
            if p:
                produk_count[p.nama] += item['qty']
    terlaris = sorted(produk_count.items(), key=lambda x: x[1], reverse=True)[:5]
    return jsonify({
        'per_hari': [{'tanggal': k, 'total': v} for k, v in list(per_hari.items())[-14:]],
        'per_bulan': [{'bulan': k, 'total': v} for k, v in per_bulan.items()],
        'terlaris': [{'nama': k, 'qty': v} for k, v in terlaris],
        'total_pendapatan': sum(t.total for t in transaksi),
        'rata_per_hari': int(sum(t.total for t in transaksi) / max(len(per_hari), 1)),
        'total_transaksi': len(transaksi),
    })

@app.route('/api/neraca')
@login_required
def neraca():
    total_pendapatan = db.session.query(db.func.sum(Transaksi.total)).scalar() or 0
    total_hpp = 0
    transaksi = Transaksi.query.all()
    for t in transaksi:
        items = json.loads(t.items or '[]')
        for item in items:
            p = Produk.query.get(item['id'])
            if p:
                total_hpp += p.modal * item['qty']
    laba_kotor = total_pendapatan - total_hpp
    nilai_stok = db.session.query(db.func.sum(Produk.modal * Produk.stok)).scalar() or 0
    piutang = db.session.query(db.func.sum(Invoice.total)).filter_by(status='Belum Bayar').scalar() or 0
    total_aset = nilai_stok + piutang + total_pendapatan
    invoice_lunas = db.session.query(db.func.sum(Invoice.total)).filter_by(status='Lunas').scalar() or 0
    return jsonify({
        'pendapatan': total_pendapatan, 'hpp': total_hpp, 'laba_kotor': laba_kotor,
        'margin': round(laba_kotor / total_pendapatan * 100, 1) if total_pendapatan > 0 else 0,
        'nilai_stok': nilai_stok, 'piutang': piutang, 'total_aset': total_aset,
        'invoice_lunas': invoice_lunas, 'total_transaksi': len(transaksi),
    })

@app.route('/api/export/excel')
@login_required
def export_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import io

    wb = Workbook()
    
    # Sheet 1: Transaksi
    ws1 = wb.active
    ws1.title = 'Transaksi'
    header_fill = PatternFill(start_color='1e3a8a', end_color='1e3a8a', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    
    headers = ['No', 'Tanggal', 'Pelanggan', 'Kasir', 'Subtotal', 'PPN', 'Total']
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    transaksi = Transaksi.query.order_by(Transaksi.tanggal.desc()).all()
    for row, t in enumerate(transaksi, 2):
        ws1.cell(row=row, column=1, value=row-1)
        ws1.cell(row=row, column=2, value=t.tanggal.strftime('%d/%m/%Y %H:%M'))
        ws1.cell(row=row, column=3, value=t.pelanggan)
        ws1.cell(row=row, column=4, value=t.kasir)
        ws1.cell(row=row, column=5, value=t.subtotal if t.subtotal is not None else t.total)
        ws1.cell(row=row, column=6, value=t.ppn or 0)
        ws1.cell(row=row, column=7, value=t.total)
    
    # Baris akumulasi transaksi
    total_row = len(transaksi) + 2
    ws1.cell(row=total_row, column=1, value='')
    ws1.cell(row=total_row, column=2, value='')
    ws1.cell(row=total_row, column=3, value='')
    ws1.cell(row=total_row, column=4, value='TOTAL')
    ws1.cell(row=total_row, column=5, value=sum(t.subtotal if t.subtotal is not None else t.total for t in transaksi))
    ws1.cell(row=total_row, column=6, value=sum(t.ppn or 0 for t in transaksi))
    ws1.cell(row=total_row, column=7, value=sum(t.total for t in transaksi))
    # Format bold untuk baris total
    for col in range(1, 8):
        ws1.cell(row=total_row, column=col).font = Font(bold=True)
        ws1.cell(row=total_row, column=col).fill = PatternFill(start_color='dbeafe', end_color='dbeafe', fill_type='solid')    
    # Auto width
    for col in range(1, 8):
        ws1.column_dimensions[get_column_letter(col)].width = 18

    # Sheet 2: Invoice
    ws2 = wb.create_sheet('Invoice')
    headers2 = ['No', 'Nomor', 'Tanggal', 'Pelanggan', 'Total', 'Status']
    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    invoices = Invoice.query.order_by(Invoice.tanggal.desc()).all()
    for row, i in enumerate(invoices, 2):
        ws2.cell(row=row, column=1, value=row-1)
        ws2.cell(row=row, column=2, value=i.nomor)
        ws2.cell(row=row, column=3, value=i.tanggal.strftime('%d/%m/%Y'))
        ws2.cell(row=row, column=4, value=i.pelanggan)
        ws2.cell(row=row, column=5, value=i.total)
        ws2.cell(row=row, column=6, value=i.status)
    
    for col in range(1, 7):
        ws2.column_dimensions[get_column_letter(col)].width = 18

    # Sheet 3: Produk
    ws3 = wb.create_sheet('Produk')
    headers3 = ['No', 'Nama Produk', 'Kategori', 'Harga Jual', 'Modal', 'Stok', 'Nilai Stok']
    for col, h in enumerate(headers3, 1):
        cell = ws3.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    produk = Produk.query.order_by(Produk.nama).all()
    for row, p in enumerate(produk, 2):
        ws3.cell(row=row, column=1, value=row-1)
        ws3.cell(row=row, column=2, value=p.nama)
        ws3.cell(row=row, column=3, value=p.kategori or 'Umum')
        ws3.cell(row=row, column=4, value=p.harga)
        ws3.cell(row=row, column=5, value=p.modal)
        ws3.cell(row=row, column=6, value=p.stok)
        ws3.cell(row=row, column=7, value=p.modal * p.stok)
    
    for col in range(1, 8):
        ws3.column_dimensions[get_column_letter(col)].width = 18

    # Sheet 4: Jurnal
    ws4 = wb.create_sheet('Jurnal')
    headers4 = ['No', 'Tanggal', 'Keterangan', 'Kategori', 'Jenis', 'Debit', 'Kredit', 'Dibuat Oleh']
    for col, h in enumerate(headers4, 1):
        cell = ws4.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    jurnal = Jurnal.query.order_by(Jurnal.tanggal.desc()).all()
    for row, j in enumerate(jurnal, 2):
        ws4.cell(row=row, column=1, value=row-1)
        ws4.cell(row=row, column=2, value=j.tanggal.strftime('%d/%m/%Y %H:%M'))
        ws4.cell(row=row, column=3, value=j.keterangan)
        ws4.cell(row=row, column=4, value=j.kategori)
        ws4.cell(row=row, column=5, value=j.jenis)
        ws4.cell(row=row, column=6, value=j.debit)
        ws4.cell(row=row, column=7, value=j.kredit)
        ws4.cell(row=row, column=8, value=j.dibuat_oleh)
    # Baris akumulasi jurnal
    jurnal_row = len(jurnal) + 2
    ws4.cell(row=jurnal_row, column=1, value='')
    ws4.cell(row=jurnal_row, column=2, value='')
    ws4.cell(row=jurnal_row, column=3, value='')
    ws4.cell(row=jurnal_row, column=4, value='')
    ws4.cell(row=jurnal_row, column=5, value='TOTAL')
    ws4.cell(row=jurnal_row, column=6, value=sum(j.debit for j in jurnal))
    ws4.cell(row=jurnal_row, column=7, value=sum(j.kredit for j in jurnal))
    ws4.cell(row=jurnal_row, column=8, value='')
    # Baris saldo
    saldo_row = jurnal_row + 1
    ws4.cell(row=saldo_row, column=5, value='SALDO')
    ws4.cell(row=saldo_row, column=6, value=sum(j.debit for j in jurnal) - sum(j.kredit for j in jurnal))
    # Format bold
    for col in range(1, 9):
        ws4.cell(row=jurnal_row, column=col).font = Font(bold=True)
        ws4.cell(row=jurnal_row, column=col).fill = PatternFill(start_color='dbeafe', end_color='dbeafe', fill_type='solid')
        ws4.cell(row=saldo_row, column=col).font = Font(bold=True)
        ws4.cell(row=saldo_row, column=col).fill = PatternFill(start_color='d1fae5', end_color='d1fae5', fill_type='solid')
    for col in range(1, 9):
        ws4.column_dimensions[get_column_letter(col)].width = 18

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return send_file(buffer, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', download_name='laporan-umkm-pro.xlsx')

@app.route('/api/export/pdf')
@login_required
def export_pdf():
    from reportlab.platypus import HRFlowable
    toko = Toko.query.first()
    nama_toko = toko.nama if toko else 'UMKM Pro'
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
        rightMargin=2*cm, leftMargin=2*cm,
        topMargin=2*cm, bottomMargin=2*cm)
    elemen = []
    
    s_judul = ParagraphStyle('judul', fontSize=16, fontName='Helvetica-Bold', alignment=1, spaceAfter=8)
    s_sub = ParagraphStyle('sub', fontSize=10, fontName='Helvetica', alignment=1, textColor=colors.HexColor('#666666'), spaceAfter=16)
    s_section = ParagraphStyle('section', fontSize=12, fontName='Helvetica-Bold', spaceAfter=8, textColor=colors.HexColor('#1e3a8a'))
    s_normal = ParagraphStyle('normal', fontSize=9, fontName='Helvetica', spaceAfter=4)

    elemen.append(Paragraph(f'LAPORAN KEUANGAN', s_judul))
    elemen.append(Paragraph(f'{nama_toko}', s_sub))
    elemen.append(Paragraph(f'Dicetak: {datetime.now().strftime("%d/%m/%Y %H:%M")}', s_sub))
    elemen.append(HRFlowable(width='100%', thickness=1.5, color=colors.HexColor('#1e3a8a'), spaceAfter=12))

    # Ringkasan
    total_pendapatan = db.session.query(db.func.sum(Transaksi.total)).scalar() or 0
    total_transaksi = Transaksi.query.count()
    total_invoice = Invoice.query.count()
    invoice_lunas = Invoice.query.filter_by(status='Lunas').count()

    elemen.append(Paragraph('RINGKASAN KEUANGAN', s_section))
    total_subtotal = db.session.query(db.func.sum(Transaksi.subtotal)).scalar() or 0
    total_ppn = db.session.query(db.func.sum(Transaksi.ppn)).scalar() or 0
    total_debit = db.session.query(db.func.sum(Jurnal.debit)).scalar() or 0
    total_kredit = db.session.query(db.func.sum(Jurnal.kredit)).scalar() or 0
    saldo_jurnal = total_debit - total_kredit
    ringkasan = [
        ['Keterangan', 'Nilai'],
        ['Total Subtotal', f'Rp {total_subtotal:,}'],
        ['Total PPN', f'Rp {total_ppn:,}'],
        ['Total Pendapatan (inc. PPN)', f'Rp {total_pendapatan:,}'],
        ['Total Transaksi', str(total_transaksi)],
        ['Total Invoice', str(total_invoice)],
        ['Invoice Lunas', str(invoice_lunas)],
        ['Invoice Belum Bayar', str(total_invoice - invoice_lunas)],
        ['Total Pemasukan Jurnal', f'Rp {total_debit:,}'],
        ['Total Pengeluaran Jurnal', f'Rp {total_kredit:,}'],
        ['Saldo Jurnal', f'Rp {saldo_jurnal:,}'],
    ]
    t_ringkasan = Table(ringkasan, colWidths=[10*cm, 7*cm])
    t_ringkasan.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1e3a8a')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 9),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#e2e8f0')),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#f0f4ff')]),
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ('LEFTPADDING', (0,0), (-1,-1), 8),
    ]))
    elemen.append(t_ringkasan)
    elemen.append(Spacer(1, 0.5*cm))

    # Transaksi terakhir
    elemen.append(Paragraph('TRANSAKSI TERAKHIR (20 Data)', s_section))
    trx_data = [['No', 'Tanggal', 'Pelanggan', 'Kasir', 'Total (Rp)']]
    transaksi = Transaksi.query.order_by(Transaksi.tanggal.desc()).limit(20).all()
    for idx, t in enumerate(transaksi, 1):
        trx_data.append([str(idx), t.tanggal.strftime('%d/%m/%Y'), t.pelanggan, t.kasir or '', f'{t.total:,}'])

    t_trx = Table(trx_data, colWidths=[1*cm, 3.5*cm, 4*cm, 3.5*cm, 5*cm])
    t_trx.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1e3a8a')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 8),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('ALIGN', (1,0), (1,-1), 'LEFT'),
        ('ALIGN', (2,0), (2,-1), 'LEFT'),
        ('ALIGN', (3,0), (3,-1), 'LEFT'),
        ('ALIGN', (4,0), (4,-1), 'RIGHT'),
        ('RIGHTPADDING', (4,0), (4,-1), 10),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#e2e8f0')),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#f0f4ff')]),
        ('TOPPADDING', (0,0), (-1,-1), 5),
        ('BOTTOMPADDING', (0,0), (-1,-1), 5),
    ]))
    elemen.append(t_trx)

    doc.build(elemen)
    buffer.seek(0)
    return send_file(buffer, mimetype='application/pdf', download_name='laporan-umkm-pro.pdf')

@app.route('/api/dashboard', methods=['GET'])
@login_required
def dashboard():
    total_pendapatan = db.session.query(db.func.sum(Transaksi.total)).scalar() or 0
    total_transaksi = Transaksi.query.count()
    invoice_belum = Invoice.query.filter_by(status='Belum Bayar').count()
    stok_menipis = Produk.query.filter(Produk.stok < 10).count()
    transaksi_terakhir = Transaksi.query.order_by(Transaksi.tanggal.desc()).limit(5).all()
    return jsonify({
        'total_pendapatan': total_pendapatan, 'total_transaksi': total_transaksi,
        'invoice_belum': invoice_belum, 'stok_menipis': stok_menipis,
        'transaksi_terakhir': [{'id': t.id, 'pelanggan': t.pelanggan, 'kasir': t.kasir, 'total': t.total, 'items': json.loads(t.items or '[]'), 'tanggal': t.tanggal.strftime('%d/%m/%Y %H:%M')} for t in transaksi_terakhir]
    })

with app.app_context():
    try:
        db.create_all()
    except Exception as e:
        print(f"Database init error: {e}")



@app.route('/api/neraca/pdf')
@login_required
def neraca_pdf():
    from reportlab.platypus import HRFlowable
    toko = Toko.query.first()
    nama_toko = toko.nama if toko else 'UMKM Pro'
    alamat_toko = toko.alamat if toko else ''
    telepon_toko = toko.telepon if toko else ''

    total_pendapatan = db.session.query(db.func.sum(Transaksi.total)).scalar() or 0
    total_hpp = 0
    transaksi_all = Transaksi.query.all()
    for t in transaksi_all:
        items = json.loads(t.items or '[]')
        for item in items:
            p = Produk.query.get(item['id'])
            if p:
                total_hpp += p.modal * item['qty']
    laba_kotor = total_pendapatan - total_hpp
    margin = round(laba_kotor / total_pendapatan * 100, 1) if total_pendapatan > 0 else 0
    nilai_stok = db.session.query(db.func.sum(Produk.modal * Produk.stok)).scalar() or 0
    piutang = db.session.query(db.func.sum(Invoice.total)).filter_by(status='Belum Bayar').scalar() or 0
    invoice_lunas = db.session.query(db.func.sum(Invoice.total)).filter_by(status='Lunas').scalar() or 0
    total_aset = nilai_stok + piutang + total_pendapatan
    total_transaksi = len(transaksi_all)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
        rightMargin=2*cm, leftMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)
    elemen = []

    s_judul = ParagraphStyle('judul', fontSize=18, fontName='Helvetica-Bold', alignment=1, spaceAfter=4)
    s_sub = ParagraphStyle('sub', fontSize=10, fontName='Helvetica', alignment=1, textColor=colors.HexColor('#666666'), spaceAfter=4)
    s_section = ParagraphStyle('section', fontSize=11, fontName='Helvetica-Bold', spaceAfter=8, textColor=colors.HexColor('#1e3a8a'))
    s_footer = ParagraphStyle('footer', fontSize=8, fontName='Helvetica', textColor=colors.HexColor('#999999'), alignment=1)

    elemen.append(Paragraph('NERACA KEUANGAN', s_judul))
    elemen.append(Paragraph(nama_toko, s_sub))
    if alamat_toko:
        elemen.append(Paragraph(alamat_toko, s_sub))
    elemen.append(Paragraph(f'Per {datetime.now().strftime("%d %B %Y")}', s_sub))
    elemen.append(HRFlowable(width='100%', thickness=1.5, color=colors.HexColor('#1e3a8a'), spaceAfter=14))

    elemen.append(Paragraph('LAPORAN LABA RUGI', s_section))
    laba_data = [
        ['Keterangan', 'Nilai'],
        ['Total Pendapatan (termasuk PPN)', f'Rp {total_pendapatan:,}'],
        ['Harga Pokok Penjualan (HPP)', f'(Rp {total_hpp:,})'],
        ['Laba Kotor', f'Rp {laba_kotor:,}'],
        ['Margin Keuntungan', f'{margin}%'],
        ['Total Transaksi', f'{total_transaksi} transaksi'],
    ]
    t_laba = Table(laba_data, colWidths=[12*cm, 5*cm])
    t_laba.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1e3a8a')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 9),
        ('ALIGN', (1,0), (1,-1), 'RIGHT'),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#e2e8f0')),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#f0f4ff')]),
        ('FONTNAME', (0,3), (-1,3), 'Helvetica-Bold'),
        ('TOPPADDING', (0,0), (-1,-1), 7),
        ('BOTTOMPADDING', (0,0), (-1,-1), 7),
        ('LEFTPADDING', (0,0), (-1,-1), 10),
        ('RIGHTPADDING', (1,0), (1,-1), 10),
    ]))
    elemen.append(t_laba)
    elemen.append(Spacer(1, 0.5*cm))

    elemen.append(Paragraph('POSISI ASET', s_section))
    aset_data = [
        ['Keterangan', 'Nilai'],
        ['Nilai Stok (Modal di Gudang)', f'Rp {nilai_stok:,}'],
        ['Piutang (Invoice Belum Bayar)', f'Rp {piutang:,}'],
        ['Invoice Lunas (Pembayaran Diterima)', f'Rp {invoice_lunas:,}'],
        ['Total Pendapatan Diterima', f'Rp {total_pendapatan:,}'],
        ['Total Aset (Estimasi)', f'Rp {total_aset:,}'],
    ]
    t_aset = Table(aset_data, colWidths=[12*cm, 5*cm])
    t_aset.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#059669')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 9),
        ('ALIGN', (1,0), (1,-1), 'RIGHT'),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#e2e8f0')),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#f0fff4')]),
        ('FONTNAME', (0,5), (-1,5), 'Helvetica-Bold'),
        ('TOPPADDING', (0,0), (-1,-1), 7),
        ('BOTTOMPADDING', (0,0), (-1,-1), 7),
        ('LEFTPADDING', (0,0), (-1,-1), 10),
        ('RIGHTPADDING', (1,0), (1,-1), 10),
    ]))
    elemen.append(t_aset)
    elemen.append(Spacer(1, 1*cm))
    elemen.append(HRFlowable(width='100%', thickness=0.5, color=colors.HexColor('#cccccc'), spaceAfter=8))
    elemen.append(Paragraph(f'Dicetak: {datetime.now().strftime("%d/%m/%Y %H:%M")}  ·  {nama_toko}' + (f'  ·  {telepon_toko}' if telepon_toko else ''), s_footer))

    doc.build(elemen)
    buffer.seek(0)
    return send_file(buffer, mimetype='application/pdf', download_name=f'neraca-{datetime.now().strftime("%Y%m%d")}.pdf')

@app.route('/api/jurnal/pdf')
@login_required
def jurnal_pdf():
    from reportlab.platypus import HRFlowable
    dari = request.args.get('dari', '')
    sampai = request.args.get('sampai', '')
    query = Jurnal.query
    if dari:
        query = query.filter(Jurnal.tanggal >= dari)
    if sampai:
        query = query.filter(Jurnal.tanggal <= sampai + ' 23:59:59')
    jurnal = query.order_by(Jurnal.tanggal.asc()).all()
    total_debit = sum(j.debit for j in jurnal)
    total_kredit = sum(j.kredit for j in jurnal)
    saldo = total_debit - total_kredit

    toko = Toko.query.first()
    nama_toko = toko.nama if toko else 'UMKM Pro'
    telepon_toko = toko.telepon if toko else ''

    if dari and sampai:
        periode = f'{dari} s/d {sampai}'
    elif dari:
        periode = f'Dari {dari}'
    elif sampai:
        periode = f'Sampai {sampai}'
    else:
        periode = 'Semua periode'

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
        rightMargin=2*cm, leftMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)
    elemen = []

    s_judul = ParagraphStyle('judul', fontSize=16, fontName='Helvetica-Bold', alignment=1, spaceAfter=4)
    s_sub = ParagraphStyle('sub', fontSize=10, fontName='Helvetica', alignment=1, textColor=colors.HexColor('#666666'), spaceAfter=4)
    s_section = ParagraphStyle('section', fontSize=11, fontName='Helvetica-Bold', spaceAfter=8, textColor=colors.HexColor('#4f46e5'))
    s_normal = ParagraphStyle('normal', fontSize=9, fontName='Helvetica', spaceAfter=4)
    s_footer = ParagraphStyle('footer', fontSize=8, fontName='Helvetica', textColor=colors.HexColor('#999999'), alignment=1)

    elemen.append(Paragraph('JURNAL KEUANGAN', s_judul))
    elemen.append(Paragraph(nama_toko, s_sub))
    elemen.append(Paragraph(f'Periode: {periode}', s_sub))
    elemen.append(Paragraph(f'Dicetak: {datetime.now().strftime("%d/%m/%Y %H:%M")}', s_sub))
    elemen.append(HRFlowable(width='100%', thickness=1.5, color=colors.HexColor('#4f46e5'), spaceAfter=12))

    ringkasan_data = [
        ['Total Pemasukan', 'Total Pengeluaran', 'Saldo'],
        [f'Rp {total_debit:,}', f'Rp {total_kredit:,}', f'Rp {saldo:,}'],
    ]
    t_ring = Table(ringkasan_data, colWidths=[6*cm, 6*cm, 5*cm])
    t_ring.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#4f46e5')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTNAME', (0,1), (-1,1), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 9),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#e2e8f0')),
        ('BACKGROUND', (0,1), (-1,1), colors.HexColor('#f5f3ff')),
        ('TOPPADDING', (0,0), (-1,-1), 8),
        ('BOTTOMPADDING', (0,0), (-1,-1), 8),
    ]))
    elemen.append(t_ring)
    elemen.append(Spacer(1, 0.5*cm))

    if jurnal:
        elemen.append(Paragraph(f'DETAIL JURNAL ({len(jurnal)} Entri)', s_section))
        data = [['No', 'Tanggal', 'Keterangan', 'Kategori', 'Pemasukan', 'Pengeluaran']]
        for idx, j in enumerate(jurnal, 1):
            debit_str = f'Rp {j.debit:,}' if j.debit > 0 else '-'
            kredit_str = f'Rp {j.kredit:,}' if j.kredit > 0 else '-'
            data.append([str(idx), j.tanggal.strftime('%d/%m/%Y'), j.keterangan[:40], j.kategori, debit_str, kredit_str])
        data.append(['', '', '', 'TOTAL', f'Rp {total_debit:,}', f'Rp {total_kredit:,}'])
        t_jurnal = Table(data, colWidths=[0.8*cm, 2.5*cm, 5.5*cm, 2.5*cm, 3*cm, 3*cm])
        t_jurnal.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#4f46e5')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,-1), 8),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('ALIGN', (2,0), (2,-1), 'LEFT'),
            ('ALIGN', (3,0), (3,-1), 'LEFT'),
            ('ALIGN', (4,0), (-1,-1), 'RIGHT'),
            ('GRID', (0,0), (-1,-1), 0.3, colors.HexColor('#e2e8f0')),
            ('ROWBACKGROUNDS', (0,1), (-1,-2), [colors.white, colors.HexColor('#f9f8ff')]),
            ('BACKGROUND', (0,-1), (-1,-1), colors.HexColor('#ede9fe')),
            ('FONTNAME', (0,-1), (-1,-1), 'Helvetica-Bold'),
            ('TOPPADDING', (0,0), (-1,-1), 5),
            ('BOTTOMPADDING', (0,0), (-1,-1), 5),
            ('LEFTPADDING', (0,0), (-1,-1), 6),
            ('RIGHTPADDING', (0,0), (-1,-1), 6),
        ]))
        elemen.append(t_jurnal)
    else:
        elemen.append(Paragraph('Tidak ada data jurnal untuk periode ini.', s_normal))

    elemen.append(Spacer(1, 0.8*cm))
    elemen.append(HRFlowable(width='100%', thickness=0.5, color=colors.HexColor('#cccccc'), spaceAfter=8))
    elemen.append(Paragraph(f'Dicetak: {datetime.now().strftime("%d/%m/%Y %H:%M")}  ·  {nama_toko}' + (f'  ·  {telepon_toko}' if telepon_toko else ''), s_footer))

    doc.build(elemen)
    buffer.seek(0)
    return send_file(buffer, mimetype='application/pdf', download_name=f'jurnal-{datetime.now().strftime("%Y%m%d")}.pdf')

@app.route('/api/jurnal', methods=['GET'])
@login_required
def get_jurnal():
    tanggal_dari = request.args.get('dari', '')
    tanggal_sampai = request.args.get('sampai', '')
    query = Jurnal.query
    if tanggal_dari:
        query = query.filter(Jurnal.tanggal >= tanggal_dari)
    if tanggal_sampai:
        query = query.filter(Jurnal.tanggal <= tanggal_sampai + ' 23:59:59')
    jurnal = query.order_by(Jurnal.tanggal.desc()).all()
    total_debit = sum(j.debit for j in jurnal)
    total_kredit = sum(j.kredit for j in jurnal)
    return jsonify({
        'jurnal': [{'id': j.id, 'tanggal': j.tanggal.strftime('%d/%m/%Y %H:%M'), 'keterangan': j.keterangan, 'kategori': j.kategori, 'jenis': j.jenis, 'debit': j.debit, 'kredit': j.kredit, 'sumber': j.sumber, 'dibuat_oleh': j.dibuat_oleh} for j in jurnal],
        'total_debit': total_debit,
        'total_kredit': total_kredit,
        'saldo': total_debit - total_kredit
    })

@app.route('/api/jurnal', methods=['POST'])
@login_required
def add_jurnal():
    b = request.json
    jenis = b.get('jenis', 'pemasukan')
    nominal = int(b.get('nominal', 0))
    j = Jurnal(
        keterangan=b['keterangan'],
        kategori=b.get('kategori', 'Umum'),
        jenis=jenis,
        debit=nominal if jenis == 'pemasukan' else 0,
        kredit=nominal if jenis == 'pengeluaran' else 0,
        sumber='manual',
        dibuat_oleh=current_user.nama or current_user.username
    )
    db.session.add(j)
    db.session.commit()
    return jsonify({'id': j.id, 'ok': True})

@app.route('/api/jurnal/<int:jid>', methods=['DELETE'])
@login_required
def del_jurnal(jid):
    if current_user.role != 'owner':
        return jsonify({'error': 'Akses ditolak'}), 403
    j = Jurnal.query.get_or_404(jid)
    db.session.delete(j)
    db.session.commit()
    return jsonify({'ok': True})

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(debug=False, host='0.0.0.0', port=port)
